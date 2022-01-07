# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from openpyxl import Workbook
import openpyxl
import string
from openpyxl.styles.fonts import Font
import os

class ExcelWriter():

    def __init__(self):
        #self.wb = openpyxl.load_workbook('/home/zorin/vinproject/vinproject/specs.xlsx')
        #self.specs_sheet = self.wb['specs']
        self.check_file = os.path.isfile('Specs.xlsx')
        # use existing file
        if self.check_file:
            self.wb = openpyxl.load_workbook('Specs.xlsx')
            self.specs_sheet = self.wb.active
            self.check = False
        # create new file if file not already exist
        else:
            self.wb = Workbook()
            self.wb['Sheet'].title = 'Specs'
            self.specs_sheet = self.wb.active
            self.check = True

    def close_spider(self,spider):
        self.wb.save('Specs.xlsx')

    def process_item(self, item, spider):
        if self.check:
            self.check = False
            self.set_labels(item)
        values = list(item.values())
        self.specs_sheet.append(values)
        return item

    # settings labels (first line) will only run first time
    def set_labels(self,item):
        labels = list(item.keys())
        alphabets = list(string.ascii_uppercase)[0:len(labels)]
        for letter,label in zip(alphabets,labels):
            self.specs_sheet[f'{letter}1'].value = label
            self.specs_sheet[f'{letter}1'].font = Font(bold=True)
